import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import calendar
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
import io
import openpyxl
from copy import copy

st.set_page_config(page_title="Hệ thống Quản lý Âu Việt Mỹ", layout="wide", page_icon="🛡️")

# ==========================================
# 1. KẾT NỐI GOOGLE SHEETS
# ==========================================
@st.cache_resource
def init_connection():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    try:
        creds_dict = json.loads(st.secrets["GOOGLE_CREDENTIALS"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)
    client = gspread.authorize(creds)
    return client.open_by_key("13Y44fuaCvd1yTZvlMzTtFoyFpfOLb-PoLTrcvkEEICY")

sheet = init_connection()

# ==========================================
# 2. CÁC HÀM XỬ LÝ DỮ LIỆU CỐT LÕI
# ==========================================
@st.cache_data(ttl=600)
def load_ds_gv():
    ds_gv = pd.DataFrame(sheet.worksheet("DS_GV").get_all_records())
    if len(ds_gv.columns) >= 5:
        ds_gv = ds_gv.rename(columns={
            ds_gv.columns[0]: 'Mã định danh',     
            ds_gv.columns[1]: 'Họ tên Giáo viên', 
            ds_gv.columns[2]: 'Tổ chuyên môn',    
            ds_gv.columns[4]: 'Mã TKB'            
        })
    
    # Bổ sung logic: Nhận diện cột "Ngày nghỉ việc" (Cột F) [cite: 18, 33, 46, 48]
    if 'Ngày nghỉ việc' not in ds_gv.columns:
        ds_gv['Ngày nghỉ việc'] = ""
        
    if 'Mã định danh' in ds_gv.columns:
        ds_gv['Mã định danh'] = ds_gv['Mã định danh'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    return ds_gv

ds_gv = load_ds_gv()

def scan_matrix_from_dataframe(df_tkb, ds_gv):
    import difflib

    unmatched_log = []
    pc_data = []

    def tach_mon_va_gv(cell_text):
        text = str(cell_text or "").strip()

        if not text or text.lower() == "nan":
            return None, None, "EMPTY"

        text = text.replace("_", "-")

        if "-" not in text:
            return None, None, "NO_SEPARATOR"

        parts = text.split("-")
        mon = parts[0].strip()
        gv_raw = parts[-1].strip()

        return mon, gv_raw, ""

    def tach_ten_gv_tu_ma(gv_raw):
        raw = str(gv_raw or "").strip()

        prefixes = [
            "t.",
            "c.",
            "mr.",
            "mrs.",
            "thầy ",
            "cô "
        ]

        matched = False
        gv_short = raw

        for p in prefixes:
            if gv_short.lower().startswith(p):
                gv_short = gv_short[len(p):].strip()
                matched = True
                break

        if not matched:
            return "", False

        # Quy tắc mới:
        # - C.Loan -> Loan
        # - T.Quốc Anh -> Quốc Anh
        # - T.Nguyễn Quốc Anh -> Quốc Anh
        words = gv_short.split()
        if len(words) >= 2:
            gv_short = " ".join(words[-2:])
        elif len(words) == 1:
            gv_short = words[0]
        else:
            gv_short = ""

        return gv_short.strip(), True

    def tao_danh_sach_ma_tkb(row):
        codes = []

        ma_tkb = str(row.get("Mã TKB", "") or "").strip()
        if ma_tkb:
            codes.extend([x.strip() for x in ma_tkb.split(",") if x.strip()])

        return codes

    def goi_y_loi_chinh_ta(gv_short, ds_gv_source):
        candidates = []

        if "Mã TKB" in ds_gv_source.columns:
            for value in ds_gv_source["Mã TKB"].astype(str):
                for code in str(value).split(","):
                    code = code.strip()
                    if code:
                        candidates.append(code)

        if "Họ tên Giáo viên" in ds_gv_source.columns:
            for ten in ds_gv_source["Họ tên Giáo viên"].astype(str):
                ten = ten.strip()
                if ten:
                    candidates.append(ten)
                    words = ten.split()
                    if len(words) >= 2:
                        candidates.append(" ".join(words[-2:]))
                    elif len(words) == 1:
                        candidates.append(words[0])

        candidates = list(set(candidates))

        matches = difflib.get_close_matches(
            str(gv_short),
            candidates,
            n=3,
            cutoff=0.72
        )

        if matches:
            return " | Gợi ý: " + ", ".join(matches)

        return ""

    header_idx = -1

    for i in range(min(15, df_tkb.shape[0])):
        row_str = " ".join([str(x).lower() for x in df_tkb.iloc[i].values])

        if "thứ" in row_str and "tiết" in row_str:
            header_idx = i
            break

    if header_idx == -1:
        header_idx = 6

    classes_info = []

    for col_idx in range(2, df_tkb.shape[1]):
        val = str(df_tkb.iloc[header_idx, col_idx]).strip()

        if val and val.lower() not in [
            "thứ",
            "tiết",
            "buổi",
            "sáng",
            "chiều"
        ]:
            classes_info.append((col_idx, val))

    current_thu = "Thứ Hai"
    current_tiet = ""

    for row_idx in range(header_idx + 1, df_tkb.shape[0]):

        val_thu = str(df_tkb.iloc[row_idx, 0]).strip()

        if val_thu:
            val_thu_lower = val_thu.lower()

            if "2" in val_thu_lower or "hai" in val_thu_lower:
                current_thu = "Thứ Hai"
            elif "3" in val_thu_lower or "ba" in val_thu_lower:
                current_thu = "Thứ Ba"
            elif "4" in val_thu_lower or "tư" in val_thu_lower:
                current_thu = "Thứ Tư"
            elif "5" in val_thu_lower or "năm" in val_thu_lower:
                current_thu = "Thứ Năm"
            elif "6" in val_thu_lower or "sáu" in val_thu_lower:
                current_thu = "Thứ Sáu"
            elif "7" in val_thu_lower or "bảy" in val_thu_lower:
                current_thu = "Thứ Bảy"

        val_tiet = str(df_tkb.iloc[row_idx, 1]).replace(".0", "").strip()

        if val_tiet:
            current_tiet = val_tiet

        for col_idx, class_name in classes_info:

            cell = str(df_tkb.iloc[row_idx, col_idx]).strip()

            try:
                mon, gv_raw, status = tach_mon_va_gv(cell)

                if status == "EMPTY":
                    continue

                if status == "NO_SEPARATOR":
                    continue

                gv_short, has_teacher_prefix = tach_ten_gv_tu_ma(gv_raw)

                if not has_teacher_prefix:
                    continue

                match = pd.DataFrame()

                if "Mã TKB" in ds_gv.columns:
                    mask = ds_gv.apply(
                        lambda row: gv_short.lower() in [
                            code.strip().lower()
                            for code in tao_danh_sach_ma_tkb(row)
                        ],
                        axis=1
                    )
                    match = ds_gv[mask]

                if match.empty and "Họ tên Giáo viên" in ds_gv.columns:
                    match = ds_gv[
                        ds_gv["Họ tên Giáo viên"]
                        .astype(str)
                        .str.lower()
                        .apply(
                            lambda ten: gv_short.lower() == " ".join(ten.split()[-2:])
                            if len(ten.split()) >= 2
                            else gv_short.lower() == ten.strip().lower()
                        )
                    ]

                if not match.empty:
                    pc_data.append({
                        "Lớp": class_name,
                        "Môn học": mon,
                        "Họ tên GV": match.iloc[0]["Họ tên Giáo viên"],
                        "Mã định danh": str(match.iloc[0]["Mã định danh"]),
                        "Thứ": current_thu,
                        "Tiết": current_tiet
                    })

                else:
                    goi_y = goi_y_loi_chinh_ta(gv_short, ds_gv)

                    unmatched_log.append(
                        f"👻 Bỏ qua: {current_thu} T.{current_tiet} "
                        f"[{class_name}] - '{cell}' "
                        f"(Không có GV '{gv_short}' trong DS_GV{goi_y})"
                    )

            except Exception as e:
                print(f"[SCAN_TKB_ERROR] {repr(e)} | Cell={cell}")

                unmatched_log.append(
                    f"❌ Lỗi đọc ô TKB: {current_thu} T.{current_tiet} "
                    f"[{class_name}] - '{cell}'. "
                    "Vui lòng kiểm tra lại định dạng ô này trong file TKB."
                )

    if pc_data:
        df_pc = pd.DataFrame(pc_data).drop_duplicates()
    else:
        df_pc = pd.DataFrame(
            columns=[
                "Lớp",
                "Môn học",
                "Họ tên GV",
                "Mã định danh",
                "Thứ",
                "Tiết"
            ]
        )

    return df_pc, unmatched_log
@st.cache_data(ttl=300)
def load_flat_tkb(thang, tuan):
    """
    Fallback dữ liệu TKB cũ dạng TKB_{thang}_W{tuan}.
    Luồng vận hành chính hiện nay dùng sheet TKB_PhanCong.
    """
    columns_default = ["Lớp", "Môn học", "Họ tên GV", "Mã định danh", "Thứ", "Tiết"]

    def doc_tkb(tab_name):
        try:
            data = sheet.worksheet(tab_name).get_all_values()
            if len(data) > 1:
                return chuan_hoa_df_tkb_phang(pd.DataFrame(data[1:], columns=data[0]))
        except Exception:
            return None
        return None

    for t in range(int(tuan), 0, -1):
        df = doc_tkb(f"TKB_{int(thang)}_W{t}")
        if df is not None:
            return df

    thang_hien_tai = int(thang)
    for offset in range(1, 13):
        thang_truoc = ((thang_hien_tai - offset - 1) % 12) + 1
        for t in range(5, 0, -1):
            df = doc_tkb(f"TKB_{thang_truoc}_W{t}")
            if df is not None:
                return df

    return pd.DataFrame(columns=columns_default)


def chuan_hoa_df_tkb_phang(df):
    """Chuẩn hóa DataFrame TKB phẳng về đúng 6 cột hệ thống cần."""
    columns_default = ["Lớp", "Môn học", "Họ tên GV", "Mã định danh", "Thứ", "Tiết"]
    if df is None or df.empty:
        return pd.DataFrame(columns=columns_default)

    df2 = df.copy()
    for col in columns_default:
        if col not in df2.columns:
            df2[col] = ""
    df2 = df2[columns_default]
    for col in columns_default:
        df2[col] = df2[col].astype(str).fillna("").str.strip()
    df2["Mã định danh"] = df2["Mã định danh"].str.replace(r"\.0$", "", regex=True).str.strip()
    df2["Tiết"] = df2["Tiết"].str.replace(r"\.0$", "", regex=True).str.strip()
    return df2


def parse_date_ddmmyyyy(value):
    """Đọc ngày dạng dd/mm/yyyy, yyyy-mm-dd hoặc date/datetime."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    return None


def parse_datetime_ddmmyyyy_hhmmss(value):
    """Đọc thời điểm dạng dd/mm/yyyy HH:MM:SS; sai định dạng thì trả về datetime nhỏ nhất."""
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return datetime.min


TKB_PHANCONG_SHEET = "TKB_PhanCong"
TKB_HIEULUC_SHEET = "TKB_HieuLuc"
TKB_PHANG_COLUMNS = ["Lớp", "Môn học", "Họ tên GV", "Mã định danh", "Thứ", "Tiết"]
TKB_HIEULUC_HEADERS = ["Tháng áp dụng", "Ngày áp dụng", "Lớp", "Môn học", "Họ tên GV", "Mã định danh", "Thứ", "Tiết", "Ngày tải", "Ghi chú"]


def dinh_dang_thang_ap_dung(ngay_ap_dung):
    """Trả về chuỗi MM/YYYY từ một ngày áp dụng hợp lệ."""
    d = parse_date_ddmmyyyy(ngay_ap_dung)
    if d is None:
        return ""
    return f"{d.month:02d}/{d.year}"


def ensure_tkb_phancong_sheet():
    """Tạo/chuẩn hóa sheet TKB_PhanCong nếu chưa có."""
    headers = TKB_PHANG_COLUMNS
    try:
        ws = sheet.worksheet(TKB_PHANCONG_SHEET)
    except Exception:
        ws = sheet.add_worksheet(title=TKB_PHANCONG_SHEET, rows="1000", cols="20")
        ws.update("A1", [headers])
        return ws

    try:
        row_1 = ws.row_values(1)
        if not row_1:
            ws.update("A1", [headers])
    except Exception:
        pass
    return ws


def ensure_tkb_hieuluc_sheet():
    """Tạo/chuẩn hóa sheet TKB_HieuLuc nếu chưa có."""
    try:
        ws = sheet.worksheet(TKB_HIEULUC_SHEET)
    except Exception:
        ws = sheet.add_worksheet(title=TKB_HIEULUC_SHEET, rows="5000", cols="20")
        ws.update("A1", [TKB_HIEULUC_HEADERS])
        return ws

    try:
        row_1 = ws.row_values(1)
        if not row_1:
            ws.update("A1", [TKB_HIEULUC_HEADERS])
    except Exception:
        pass
    return ws


@st.cache_data(ttl=300)
def load_tkb_phancong():
    """Đọc TKB hiện hành từ sheet TKB_PhanCong."""
    try:
        data = sheet.worksheet(TKB_PHANCONG_SHEET).get_all_values()
        if len(data) > 1:
            return chuan_hoa_df_tkb_phang(pd.DataFrame(data[1:], columns=data[0]))
    except Exception:
        pass
    return pd.DataFrame(columns=TKB_PHANG_COLUMNS)


@st.cache_data(ttl=300)
def load_tkb_hieuluc_all():
    """Đọc toàn bộ dữ liệu TKB_HieuLuc."""
    try:
        data = sheet.worksheet(TKB_HIEULUC_SHEET).get_all_values()
        if len(data) > 1:
            df = pd.DataFrame(data[1:], columns=data[0])
            for col in TKB_HIEULUC_HEADERS:
                if col not in df.columns:
                    df[col] = ""
            df = df[TKB_HIEULUC_HEADERS]
            for col in TKB_HIEULUC_HEADERS:
                df[col] = df[col].astype(str).fillna("").str.strip()
            df["Mã định danh"] = df["Mã định danh"].str.replace(r"\.0$", "", regex=True).str.strip()
            df["Tiết"] = df["Tiết"].str.replace(r"\.0$", "", regex=True).str.strip()
            return df
    except Exception:
        pass
    return pd.DataFrame(columns=TKB_HIEULUC_HEADERS)


def append_tkb_hieuluc(df_pc, ngay_ap_dung, ghi_chu=""):
    """
    Ghi TKB vào TKB_HieuLuc.
    Quy tắc mới:
    - Nếu ngày áp dụng đã có dữ liệu: backup các dòng cũ rồi xóa, sau đó ghi bản mới.
    - Nếu ngày áp dụng chưa có: ghi mới.
    """
    ngay_hl = parse_date_ddmmyyyy(ngay_ap_dung)
    if ngay_hl is None:
        raise ValueError("Ngày áp dụng TKB không hợp lệ.")

    df_up = chuan_hoa_df_tkb_phang(df_pc).astype(str).fillna("")
    if df_up.empty:
        return 0

    ws = ensure_tkb_hieuluc_sheet()
    ngay_ap_dung_str = ngay_hl.strftime("%d/%m/%Y")
    thang_ap_dung = dinh_dang_thang_ap_dung(ngay_hl)
    ngay_tai_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Đọc dữ liệu hiện có để tìm các dòng trùng ngày áp dụng
    data_old = ws.get_all_values()
    rows_to_backup = []
    row_indexes_to_delete = []

    if len(data_old) > 1:
        headers = data_old[0]
        try:
            idx_ngay = headers.index("Ngày áp dụng")
            for row_index, row in enumerate(data_old[1:], start=2):
                if len(row) > idx_ngay and str(row[idx_ngay]).strip() == ngay_ap_dung_str:
                    rows_to_backup.append(row)
                    row_indexes_to_delete.append(row_index)
        except ValueError:
            pass

    # Backup các dòng cũ trước khi xóa
    if rows_to_backup:
        backup_sheet_name = "TKB_HieuLuc_Backup"
        try:
            ws_backup = sheet.worksheet(backup_sheet_name)
        except Exception:
            ws_backup = sheet.add_worksheet(title=backup_sheet_name, rows="5000", cols="20")
            ws_backup.update("A1", [TKB_HIEULUC_HEADERS + ["Thời điểm backup", "Lý do backup"]])

        backup_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        backup_rows = []
        for row in rows_to_backup:
            row_fixed = row[:len(TKB_HIEULUC_HEADERS)]
            while len(row_fixed) < len(TKB_HIEULUC_HEADERS):
                row_fixed.append("")
            backup_rows.append(row_fixed + [backup_time, f"Thay thế TKB cùng ngày áp dụng {ngay_ap_dung_str}"])

        ws_backup.append_rows(backup_rows, value_input_option="USER_ENTERED")

        # Xóa từ dưới lên để không lệch chỉ số dòng
        for row_index in sorted(row_indexes_to_delete, reverse=True):
            ws.delete_rows(row_index)

    # Ghi bản TKB mới
    rows = []
    for _, row in df_up.iterrows():
        rows.append([
            thang_ap_dung,
            ngay_ap_dung_str,
            row.get("Lớp", ""),
            row.get("Môn học", ""),
            row.get("Họ tên GV", ""),
            row.get("Mã định danh", ""),
            row.get("Thứ", ""),
            row.get("Tiết", ""),
            ngay_tai_str,
            ghi_chu,
        ])

    ws.append_rows(rows, value_input_option="USER_ENTERED")
    st.cache_data.clear()
    return len(rows)


def seed_tkb_phancong_to_hieuluc_if_needed(ngay_ap_dung):
    """
    Khi lần đầu dùng TKB_HieuLuc, tự lưu TKB_PhanCong hiện có làm bản nền của tháng.
    Mục đích: nếu TKB mới áp dụng giữa tháng, các ngày trước đó vẫn còn bản TKB cũ để tra cứu.
    """
    ngay_hl = parse_date_ddmmyyyy(ngay_ap_dung)
    if ngay_hl is None:
        return 0

    df_current = load_tkb_phancong()
    if df_current.empty:
        return 0

    df_hl = load_tkb_hieuluc_all()
    if not df_hl.empty:
        ngay_series = df_hl["Ngày áp dụng"].apply(parse_date_ddmmyyyy)
        if any((d is not None and d <= ngay_hl) for d in ngay_series):
            return 0

    ngay_nen = date(ngay_hl.year, ngay_hl.month, 1)
    if ngay_nen >= ngay_hl:
        return 0

    return append_tkb_hieuluc(
        df_current,
        ngay_nen,
        "Tự động lưu TKB_PhanCong hiện có làm bản nền trước khi cập nhật TKB mới."
    )


def save_tkb_to_phancong(df_pc, ngay_ap_dung=None, ghi_chu=""):
    """
    Lưu TKB đã quét.
    - Nếu không truyền ngày áp dụng: giữ hành vi cũ, ghi đè TKB_PhanCong.
    - Nếu có ngày áp dụng: lưu vào TKB_HieuLuc; chỉ cập nhật TKB_PhanCong khi ngày áp dụng không sau hôm nay.
    """
    df_up = chuan_hoa_df_tkb_phang(df_pc).astype(str).fillna("")

    if ngay_ap_dung is not None:
        ngay_hl = parse_date_ddmmyyyy(ngay_ap_dung)
        if ngay_hl is None:
            raise ValueError("Ngày áp dụng TKB không hợp lệ.")

        so_dong_nen = seed_tkb_phancong_to_hieuluc_if_needed(ngay_hl)
        so_dong_hieuluc = append_tkb_hieuluc(df_up, ngay_hl, ghi_chu)

        so_dong_phancong = 0
        if ngay_hl <= datetime.now().date():
            ws = ensure_tkb_phancong_sheet()
            ws.clear()
            data = [df_up.columns.tolist()] + df_up.values.tolist()
            ws.update("A1", data)
            so_dong_phancong = len(df_up)

        st.cache_data.clear()
        return {
            "so_dong_hieuluc": so_dong_hieuluc,
            "so_dong_phancong": so_dong_phancong,
            "so_dong_nen": so_dong_nen,
        }

    ws = ensure_tkb_phancong_sheet()
    ws.clear()
    data = [df_up.columns.tolist()] + df_up.values.tolist()
    ws.update("A1", data)
    st.cache_data.clear()
    return len(df_up)


def get_tkb_from_hieuluc_by_date(target_date):
    """Lấy TKB hiệu lực tại một ngày: ngày áp dụng gần nhất và không sau ngày cần tính."""
    target = parse_date_ddmmyyyy(target_date)
    if target is None:
        return pd.DataFrame(columns=TKB_PHANG_COLUMNS)

    df_hl = load_tkb_hieuluc_all()
    if df_hl.empty:
        return pd.DataFrame(columns=TKB_PHANG_COLUMNS)

    df_work = df_hl.copy()
    df_work["__ngay_ap_dung"] = df_work["Ngày áp dụng"].apply(parse_date_ddmmyyyy)
    df_work = df_work[df_work["__ngay_ap_dung"].notna()]
    df_work = df_work[df_work["__ngay_ap_dung"] <= target]
    if df_work.empty:
        return pd.DataFrame(columns=TKB_PHANG_COLUMNS)

    ngay_chon = max(df_work["__ngay_ap_dung"])
    df_selected = df_work[df_work["__ngay_ap_dung"] == ngay_chon].copy()
    df_selected["__ngay_tai"] = df_selected["Ngày tải"].apply(parse_datetime_ddmmyyyy_hhmmss)
    thoi_diem_tai_chon = max(df_selected["__ngay_tai"])
    df_selected = df_selected[df_selected["__ngay_tai"] == thoi_diem_tai_chon].copy()
    return chuan_hoa_df_tkb_phang(df_selected[TKB_PHANG_COLUMNS])


@st.cache_data(ttl=300)
def load_flat_tkb_by_date(target_date):
    """
    Lấy TKB theo ngày hiệu lực.
    Quy tắc: chọn bản TKB có Ngày áp dụng <= ngày cần tính và gần ngày cần tính nhất.
    Nếu TKB_HieuLuc chưa có dữ liệu phù hợp, fallback về TKB_PhanCong, sau đó fallback về TKB_{thang}_W{tuan} cũ.
    """
    target = parse_date_ddmmyyyy(target_date)
    if target is None:
        return pd.DataFrame(columns=TKB_PHANG_COLUMNS)

    df_hl = get_tkb_from_hieuluc_by_date(target)
    if not df_hl.empty:
        return df_hl

    df_current = load_tkb_phancong()
    if not df_current.empty:
        return df_current

    tuan_legacy = (target.day - 1) // 7 + 1
    return load_flat_tkb(target.month, tuan_legacy)


def build_tkb_by_date_cache(month, year):
    """Tạo cache TKB theo từng ngày trong tháng, mỗi ngày lấy đúng TKB hiệu lực của ngày đó."""
    last_day = calendar.monthrange(int(year), int(month))[1]
    cache = {}
    for day in range(1, last_day + 1):
        d = date(int(year), int(month), day)
        cache[d.strftime("%d/%m/%Y")] = load_flat_tkb_by_date(d)
    return cache


def get_month_calendar(year, month):
    cal = calendar.monthcalendar(year, month)
    weeks = []
    for week in cal:
        days = [d for d in week if d != 0]
        if days:
            start_date = f"{days[0]:02d}/{month:02d}"
            end_date = f"{days[-1]:02d}/{month:02d}"
            weeks.append({"days": week, "title": f"{start_date} - {end_date}"})
    return weeks

def extract_grade_safe(class_name):
    import re
    class_str = str(class_name)
    m = re.search(r'\d+', class_str)
    if m is not None:
        return m.group()
    else:
        return "Khác"

def normalize_dataframe_for_streamlit(df):
    """Chuẩn hóa kiểu dữ liệu trước khi đưa vào st.dataframe để tránh lỗi PyArrow.
    Đặc biệt cột 'Tiết' có thể chứa cả số tiết và giá trị chữ như 'ALL'.
    """
    if df is None:
        return df
    df_view = df.copy()
    for col in df_view.columns:
        if col in ['Tiết', 'Mã định danh', 'ID GV vắng', 'ID GV dạy thay', 'ID_Hang']:
            df_view[col] = df_view[col].astype(str)
    return df_view


def parse_ngay_nghi_viec(value):
    """Chuyển chuỗi ngày nghỉ việc dd/mm/yyyy thành date; rỗng hoặc sai định dạng thì xem như chưa nghỉ."""
    value = str(value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%d/%m/%Y").date()
    except Exception:
        return None


def tao_dict_ngay_nghi_viec(ds_gv_source):
    """Tạo dict: Mã định danh -> ngày nghỉ việc."""
    if ds_gv_source is None or ds_gv_source.empty or 'Ngày nghỉ việc' not in ds_gv_source.columns:
        return {}
    result = {}
    for _, row in ds_gv_source.iterrows():
        gv_id = str(row.get('Mã định danh', '')).strip()
        if gv_id:
            result[gv_id] = parse_ngay_nghi_viec(row.get('Ngày nghỉ việc', ''))
    return result


def gv_con_hieu_luc_trong_thang(ngay_nghi_viec, month, year):
    """GV còn được đưa vào xử lý tháng nếu chưa nghỉ hoặc ngày nghỉ việc sau ngày đầu tháng."""
    if ngay_nghi_viec is None:
        return True
    first_day = date(year, month, 1)
    return ngay_nghi_viec > first_day


def gv_duoc_tinh_o_ngay(ngay_nghi_viec, ngay_iter):
    """Chỉ tính tiết trước ngày nghỉ việc; từ ngày nghỉ việc trở đi không ghi nhận."""
    if ngay_nghi_viec is None:
        return True
    return ngay_iter < ngay_nghi_viec

# ==========================================
# 3. HÀM DÙNG CHUNG: FORM THÊM GIÁO VIÊN
# ==========================================
def form_them_giao_vien(form_key_prefix):
    st.subheader("Thêm Giáo viên mới vào hệ thống (Cập nhật Tab DS_GV)")

    with st.form(f"form_them_gv_{form_key_prefix}"):
        col_g1, col_g2 = st.columns(2)

        with col_g1:
            ma_gv = st.text_input("Mã định danh (Cột A) *", help="Ví dụ: GV099").strip()
            ten_gv = st.text_input("Họ tên Giáo viên (Cột B) *").strip()
            email_gv = st.text_input("Email (Cột D)").strip()

        with col_g2:
            to_cm = st.text_input("Tổ chuyên môn (Cột C)").strip()
            ma_tkb = st.text_input(
                "Mã TKB (Cột E) *",
                help="Nhập mã đúng như dùng trong TKB. Có thể nhập nhiều mã, cách nhau bằng dấu phẩy. Ví dụ: Quốc Anh, QAnh, Q.Anh"
            ).strip()
            ngay_nghi = st.text_input("Ngày nghỉ việc (Nếu có)", help="Định dạng: dd/mm/yyyy").strip()

        submit_btn = st.form_submit_button("💾 Lưu Giáo viên", type="primary")

        if submit_btn:
            if not ma_gv or not ten_gv or not ma_tkb:
                st.error("⚠️ Vui lòng nhập đầy đủ Mã định danh, Họ tên Giáo viên và Mã TKB.")
                return

            if ngay_nghi:
                try:
                    datetime.strptime(ngay_nghi, "%d/%m/%Y")
                except Exception:
                    st.error("⚠️ Ngày nghỉ việc chưa đúng định dạng dd/mm/yyyy.")
                    return

            if ma_gv in ds_gv['Mã định danh'].astype(str).values:
                st.error("⚠️ Mã định danh này đã tồn tại. Vui lòng kiểm tra lại DS_GV.")
                return

            mask_name = ds_gv['Họ tên Giáo viên'].astype(str).str.strip().str.lower() == ten_gv.lower()
            df_same_name = ds_gv[mask_name]

            is_duplicate = False
            loi_trung = ""

            if not df_same_name.empty:
                for _, row in df_same_name.iterrows():
                    old_to = str(row.get('Tổ chuyên môn', '')).strip().lower()
                    old_email = str(row.get('Email', '')).strip().lower()
                    old_matkb_list = [
                        x.strip().lower()
                        for x in str(row.get('Mã TKB', '')).split(',')
                        if x.strip()
                    ]

                    new_matkb_list = [
                        x.strip().lower()
                        for x in ma_tkb.split(',')
                        if x.strip()
                    ]

                    if to_cm and to_cm.lower() == old_to:
                        is_duplicate, loi_trung = True, "Tổ chuyên môn"
                        break

                    if email_gv and email_gv.lower() == old_email:
                        is_duplicate, loi_trung = True, "Email"
                        break

                    if set(new_matkb_list) & set(old_matkb_list):
                        is_duplicate, loi_trung = True, "Mã TKB"
                        break

            if is_duplicate:
                st.error(f"⚠️ Phát hiện trùng lặp: Giáo viên '{ten_gv}' có cùng {loi_trung} với hồ sơ đã tồn tại.")
                return

            with st.spinner("Đang lưu vào Google Sheets..."):
                try:
                    new_row = [ma_gv, ten_gv, to_cm, email_gv, ma_tkb, ngay_nghi]
                    sheet.worksheet("DS_GV").append_row(new_row)
                    load_ds_gv.clear()

                    st.success(f"✅ Đã thêm giáo viên **{ten_gv}** vào DS_GV.")
                    st.info(
                        "Lưu ý: Nếu giáo viên có dạy trong TKB, cần bảo đảm TKB dùng đúng Mã TKB vừa khai báo. "
                        "Nếu giáo viên có tính lương, cần kiểm tra ID giáo viên trong file lương."
                    )

                except Exception as e:
                    print(f"[ADD_GV_ERROR] {repr(e)}")
                    st.error(
                        "❌ Không lưu được giáo viên vào Google Sheets. "
                        "Vui lòng kiểm tra kết nối mạng, quyền truy cập Google Sheet của Service Account, "
                        "hoặc thử lại sau."
                    )

# ==========================================
# 4. HÀM TẠO EXCEL MẪU (CÓ TÙY CHỌN KHÓA SHEET)
# ==========================================
def tao_excel_mau_avm(gv_dict, weeks, month, year, dict_tkb_cac_tuan, df_nl_all, is_teacher=False, gv_nghi_viec_dict=None):
    try:
        wb = openpyxl.load_workbook("BaoCaoMau.xlsx")
    except FileNotFoundError:
        return None

    template_ws = wb.active
    template_ws_name = template_ws.title
    danh_sach_thu = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy"]
    today_date = datetime.now().date()
    nl_ngay_nghi = df_nl_all[df_nl_all['Loại ngoại lệ'] == 'Ngày nghỉ/Sự kiện'] if not df_nl_all.empty else pd.DataFrame()
    gv_nghi_viec_dict = gv_nghi_viec_dict or {}

    for gv_id, gv_name in gv_dict.items():
        gv_id = str(gv_id).strip()
        ngay_nghi_viec = gv_nghi_viec_dict.get(gv_id)
        if not gv_con_hieu_luc_trong_thang(ngay_nghi_viec, month, year):
            continue

        co_day_khong = False
        tkb_gv_cac_tuan = {}

        # dict_tkb_cac_tuan có thể là:
        # - Chuẩn cũ: {1: df_tkb_w1, 2: df_tkb_w2, ...}
        # - Chuẩn mới: {"dd/mm/yyyy": df_tkb_hieu_luc_cua_ngay, ...}
        for key_tkb, tkb_w in (dict_tkb_cac_tuan or {}).items():
            if tkb_w is not None and not tkb_w.empty:
                tkb_gv_w = tkb_w[tkb_w['Mã định danh'].astype(str).str.strip() == gv_id.strip()]
                tkb_gv_cac_tuan[key_tkb] = tkb_gv_w
                if not tkb_gv_w.empty:
                    co_day_khong = True
            else:
                tkb_gv_cac_tuan[key_tkb] = pd.DataFrame()

        nl_gv_v = df_nl_all[df_nl_all['ID GV vắng'].astype(str).str.strip() == gv_id.strip()] if not df_nl_all.empty else pd.DataFrame()
        nl_gv_dt = df_nl_all[df_nl_all['ID GV dạy thay'].astype(str).str.strip() == gv_id.strip()] if not df_nl_all.empty else pd.DataFrame()

        if not co_day_khong and nl_gv_v.empty and nl_gv_dt.empty: 
            continue 

        ws = wb.copy_worksheet(template_ws)
        ws.title = gv_name[:31]
        ws['J4'] = gv_name
        ws['S4'] = month
        ws['L71'] = gv_name
        ws['Y72'] = gv_id
        ws.column_dimensions['Y'].hidden = True
        ws.row_dimensions[72].hidden = True

        last_day_of_month = calendar.monthrange(year, month)[1]
        for w_idx, w in enumerate(weeks):
            if w_idx >= 5: break 
            valid_days = [d for d in w['days'][:6] if d != 0] 
            if not valid_days: continue
            
            start_str = f"{valid_days[0]:02d}/{month:02d}"
            end_str = f"{valid_days[-1]:02d}/{month:02d}"

            if w_idx == 0:
                ws['C6'] = f"01/{month:02d}"
                ws['E6'] = end_str           
            elif w_idx == 1:
                ws['F6'] = start_str
                ws['H6'] = end_str           
            elif w_idx == 2:
                ws['I6'] = start_str
                ws['K6'] = end_str           
            elif w_idx == 3:
                ws['L6'] = start_str
                ws['N6'] = end_str           
            elif w_idx == 4:
                ws['O6'] = start_str
                ws['Q6'] = f"{last_day_of_month:02d}/{month:02d}" 

        for thu in danh_sach_thu:
            thu_idx = danh_sach_thu.index(thu)
            for tiet in range(1, 9): 
                row_idx = 7 + (thu_idx * 10) + ((tiet - 1) if tiet <= 4 else tiet)
                for w_idx, w in enumerate(weeks):
                    if w_idx >= 5: break
                    col_idx = 4 + (w_idx * 3) 
                    day = w['days'][thu_idx]
                    if day != 0:
                        ngay_iter = date(year, month, day)
                        ngay_str = f"{day:02d}/{month:02d}/{year}"
                        
                        if not gv_duoc_tinh_o_ngay(ngay_nghi_viec, ngay_iter):
                            continue

                        if ngay_iter <= today_date:
                            tkb_key_ngay = ngay_str
                            tkb_tuan_nay = tkb_gv_cac_tuan.get(tkb_key_ngay, pd.DataFrame())
                            if tkb_tuan_nay.empty:
                                # Fallback cho dữ liệu cũ theo tuần.
                                tkb_tuan_nay = tkb_gv_cac_tuan.get(w_idx + 1, pd.DataFrame())

                            base_class = ""
                            if not tkb_tuan_nay.empty:
                                tkb_match = tkb_tuan_nay[(tkb_tuan_nay['Thứ'] == thu) & (tkb_tuan_nay['Tiết'] == str(tiet))]
                                if not tkb_match.empty:
                                    base_class = tkb_match.iloc[0]['Lớp']
                            
                            is_ngay_nghi = False
                            if not nl_ngay_nghi.empty:
                                match_nghi = nl_ngay_nghi[(nl_ngay_nghi['Ngày'] == ngay_str) & ((nl_ngay_nghi['Lớp'] == 'ALL') | (nl_ngay_nghi['Lớp'] == base_class))]
                                if not match_nghi.empty: 
                                    is_ngay_nghi = True
                            
                            nl_sk_match = nl_gv_v[(nl_gv_v['Ngày'] == ngay_str) & (nl_gv_v['Tiết'].astype(str) == str(tiet)) & (nl_gv_v['Loại ngoại lệ'] == 'Nghỉ Sự kiện/Thi')] if not nl_gv_v.empty else pd.DataFrame()
                            nl_v_match = nl_gv_v[(nl_gv_v['Ngày'] == ngay_str) & (nl_gv_v['Tiết'].astype(str) == str(tiet)) & (nl_gv_v['Loại ngoại lệ'] != 'Nghỉ Sự kiện/Thi')] if not nl_gv_v.empty else pd.DataFrame()
                            nl_dt_match = nl_gv_dt[(nl_gv_dt['Ngày'] == ngay_str) & (nl_gv_dt['Tiết'].astype(str) == str(tiet)) & (nl_gv_dt['Loại ngoại lệ'] != 'Dạy bù')] if not nl_gv_dt.empty else pd.DataFrame()
                            nl_db_match = nl_gv_dt[(nl_gv_dt['Ngày'] == ngay_str) & (nl_gv_dt['Tiết'].astype(str) == str(tiet)) & (nl_gv_dt['Loại ngoại lệ'] == 'Dạy bù')] if not nl_gv_dt.empty else pd.DataFrame()

                            target_cell = ws.cell(row=row_idx, column=col_idx)
                            new_font = copy(target_cell.font)
                            new_font.bold = True
                            
                            if is_ngay_nghi or not nl_sk_match.empty:
                                if base_class:
                                    target_cell.value = f"N({base_class})"
                                    new_font.color = "0070C0"
                            elif not nl_v_match.empty:
                                target_cell.value = f"V ({nl_v_match.iloc[0]['Lớp']})"
                                new_font.color = "FF0000"
                            elif not nl_db_match.empty:
                                target_cell.value = f"{nl_db_match.iloc[0]['Lớp']} (bù)"
                                new_font.color = "00B050"
                            elif not nl_dt_match.empty:
                                target_cell.value = f"{nl_dt_match.iloc[0]['Lớp']} (DT)"
                                new_font.color = "00B050"
                            else:
                                if base_class: 
                                    target_cell.value = base_class
                            target_cell.font = new_font
        
        # BẢO MẬT: KHÓA SHEET NẾU ĐỐI TƯỢNG LÀ GIÁO VIÊN
        if is_teacher:
            ws.protection.sheet = True
            ws.protection.password = 'avm123'

    if len(wb.sheetnames) > 1: 
        wb.remove(wb[template_ws_name])
    else: 
        template_ws.title = "KhongCoDuLieu"
        template_ws['A1'] = "Giáo viên không có phân công trong tháng."

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==========================================
# 4B. HÀM XUẤT BẢNG LƯƠNG THÁNG TỪ FILE CHẤM CÔNG
# ==========================================
# CẤU HÌNH FILE LƯƠNG CHUẨN
# Template cố định: cột SL THPT đầu tiên là D, mỗi cụm GV rộng 6 cột.
# ID GV nằm tại dòng 50, cột liền trước cột SL THPT.
LUONG_FIRST_SL_THPT_COL = 4   # D
LUONG_GROUP_WIDTH = 6
LUONG_ID_ROW = 50
LUONG_ROWS_TO_WRITE = [12, 13, 14, 15]

def chuan_hoa_ten_gv_avm(value):
    """Chuẩn hóa họ tên để khớp giữa file chấm công và file lương."""
    if value is None:
        return ""
    return " ".join(str(value).strip().split()).lower()


def lay_gia_tri_o_an_toan(value):
    """Chuyển giá trị trong ô Excel thành chuỗi an toàn để nhận diện lớp."""
    if value is None:
        return ""
    return str(value).strip()


def chuan_hoa_id_gv_avm(value):
    """Chuẩn hóa ID GV đọc từ Google Sheets hoặc Excel lương.
    Excel đôi khi trả ID dạng số hoặc chuỗi có .0 ở cuối.
    """
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.strip()


def dem_tiet_theo_khoi_tu_sheet_cham_cong(ws):
    """
    Đọc số tiết thực dạy theo khối từ sheet chấm công tháng.
    Dòng tổng trong mẫu là cố định ở dòng 67, nhưng openpyxl không tự tính công thức.
    Vì vậy hàm đếm trực tiếp vùng dữ liệu C7:T65, tương đương logic công thức dòng 67.
    """
    ket_qua = {
        "L6": 0,
        "L7": 0,
        "L8": 0,
        "L9": 0,
        "L10_11": 0,
        "L12": 0,
    }

    for row in range(7, 66):
        for col in range(3, 21):
            raw = lay_gia_tri_o_an_toan(ws.cell(row=row, column=col).value)
            if not raw:
                continue

            # Bỏ qua các mã không phải lớp thực dạy: V(...), N(...), ô trống hoặc ghi chú khác.
            raw_upper = raw.upper()
            if raw_upper.startswith("V") or raw_upper.startswith("N"):
                continue

            if raw.startswith("6"):
                ket_qua["L6"] += 1
            elif raw.startswith("7"):
                ket_qua["L7"] += 1
            elif raw.startswith("8"):
                ket_qua["L8"] += 1
            elif raw.startswith("9"):
                ket_qua["L9"] += 1
            elif raw.startswith("10") or raw.startswith("11"):
                ket_qua["L10_11"] += 1
            elif raw.startswith("12"):
                ket_qua["L12"] += 1

    return ket_qua


def tao_dict_tiet_luong_tu_excel_cham_cong(excel_cham_cong_bytes):
    """Tạo dict: Mã định danh GV -> số tiết theo 4 dòng cần ghi vào bảng lương."""
    wb_cc = openpyxl.load_workbook(io.BytesIO(excel_cham_cong_bytes), data_only=False)
    dict_tiet = {}
    ten_gv_goc = {}

    for ws in wb_cc.worksheets:
        if ws.title == "KhongCoDuLieu":
            continue

        gv_id = chuan_hoa_id_gv_avm(ws["Y72"].value)
        gv_name = ws["J4"].value if ws["J4"].value else ws.title

        # Fallback an toàn cho file chấm công cũ chưa có ô kỹ thuật Y72.
        # Phần lương chuẩn vẫn khớp bằng ID dòng 50 của file lương.
        if not gv_id:
            gv_id = chuan_hoa_id_gv_avm(ws.title)

        if not gv_id:
            continue

        theo_khoi = dem_tiet_theo_khoi_tu_sheet_cham_cong(ws)
        dict_tiet[gv_id] = {
            12: theo_khoi["L6"] + theo_khoi["L7"] + theo_khoi["L8"],
            13: theo_khoi["L9"],
            14: theo_khoi["L10_11"],
            15: theo_khoi["L12"],
        }
        ten_gv_goc[gv_id] = " ".join(str(gv_name).strip().split())

    return dict_tiet, ten_gv_goc


def tim_cac_cum_gv_trong_sheet_luong(ws):
    """
    Tìm các cụm GV trong sheet lương bằng địa chỉ cột cố định, không dò theo text header.

    Chuẩn file TH_luong.xlsx:
    - Cột SL THPT đầu tiên: D
    - Mỗi cụm GV rộng 6 cột
    - ID GV: dòng 50, cột liền trước SL THPT
    - Ghi số tiết: dòng 12, 13, 14, 15 tại cột SL THPT
    """
    ds_cum = []
    for col_sl in range(LUONG_FIRST_SL_THPT_COL, ws.max_column + 1, LUONG_GROUP_WIDTH):
        start_col = col_sl - 1
        gv_name = ws.cell(row=3, column=start_col).value
        gv_id = chuan_hoa_id_gv_avm(ws.cell(row=LUONG_ID_ROW, column=start_col).value)

        # Vẫn đưa cụm vào danh sách để cảnh báo rõ cụm nào thiếu ID,
        # nhưng không dùng tên/header để quyết định cột ghi.
        ds_cum.append({
            "ten_gv": " ".join(str(gv_name or "").strip().split()),
            "gv_id": gv_id,
            "cot_bat_dau": start_col,
            "cot_sl_thpt": col_sl,
        })
    return ds_cum


def tao_bang_luong_thang_avm(month, year, dict_tkb_cac_tuan, df_nl_all, ds_gv_source):
    """
    Tạo file TH_luong.xlsx đã điền số tiết thực dạy vào cột SL THPT.
    Chỉ ghi vào hai sheet: B. GV- NVCH và D. GV TT .
    """
    template_luong_path = "TH_luong.xlsx"
    try:
        wb_luong = openpyxl.load_workbook(template_luong_path)
    except FileNotFoundError:
        return None, [f"Không tìm thấy file mẫu {template_luong_path} trong thư mục app."], []

    weeks = get_month_calendar(year, month)
    gv_nghi_viec_dict = tao_dict_ngay_nghi_viec(ds_gv_source)
    gv_dict_all = {
        str(row["Mã định danh"]).strip(): str(row["Họ tên Giáo viên"]).strip()
        for _, row in ds_gv_source.iterrows()
        if str(row.get("Mã định danh", "")).strip()
        and str(row.get("Họ tên Giáo viên", "")).strip()
        and gv_con_hieu_luc_trong_thang(gv_nghi_viec_dict.get(str(row.get("Mã định danh", "")).strip()), month, year)
    }

    excel_cham_cong = tao_excel_mau_avm(
        gv_dict_all,
        weeks,
        month,
        year,
        dict_tkb_cac_tuan,
        df_nl_all,
        is_teacher=False,
        gv_nghi_viec_dict=gv_nghi_viec_dict
    )

    if not excel_cham_cong:
        return None, ["Không tạo được file chấm công tháng để làm nguồn điền bảng lương."], []

    dict_tiet, ten_gv_cham_cong = tao_dict_tiet_luong_tu_excel_cham_cong(excel_cham_cong)
    sheets_luong_can_dien = ["B. GV- NVCH", "D. GV TT "]
    da_dien = set()
    khong_khop_trong_luong = []

    for sheet_name in sheets_luong_can_dien:
        if sheet_name not in wb_luong.sheetnames:
            khong_khop_trong_luong.append(f"Thiếu sheet lương: {sheet_name}")
            continue

        ws_luong = wb_luong[sheet_name]
        ds_cum = tim_cac_cum_gv_trong_sheet_luong(ws_luong)

        for cum in ds_cum:
            gv_id = cum["gv_id"]
            ten_hien_thi = cum["ten_gv"] or f"Cột {cum['cot_bat_dau']}"
            if not gv_id:
                khong_khop_trong_luong.append(f"{sheet_name}: {ten_hien_thi} - thiếu ID dòng 50")
                continue
            if gv_id not in dict_tiet:
                khong_khop_trong_luong.append(f"{sheet_name}: {ten_hien_thi} - ID {gv_id} chưa có dữ liệu chấm công")
                continue

            col_sl = cum["cot_sl_thpt"]
            for row_idx in LUONG_ROWS_TO_WRITE:
                ws_luong.cell(row=row_idx, column=col_sl).value = dict_tiet[gv_id].get(row_idx, 0)
            da_dien.add(gv_id)

    khong_co_trong_file_luong = []
    for key, ten in ten_gv_cham_cong.items():
        if key not in da_dien:
            khong_co_trong_file_luong.append(ten)

    try:
        wb_luong.calculation.fullCalcOnLoad = True
        wb_luong.calculation.forceFullCalc = True
    except Exception:
        pass

    output = io.BytesIO()
    wb_luong.save(output)
    return output.getvalue(), khong_khop_trong_luong, khong_co_trong_file_luong

# ==========================================
# 5. MÀN HÌNH ĐĂNG NHẬP & PHÂN QUYỀN
# ==========================================
if "logged_in" not in st.session_state:
    st.session_state.update({"logged_in": False, "role": None, "user_name": None, "user_id": None})

if not st.session_state.logged_in:
    st.markdown("<h1 style='text-align: center;'>🛡️ CỔNG ĐĂNG NHẬP ÂU VIỆT MỸ</h1>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        loai_tk = st.selectbox("Vai trò của bạn:", ["Giáo viên", "Giám thị", "Ban Giám Hiệu"])
        mat_khau = st.text_input("Mật khẩu / Mã định danh:", type="password")
      
        if st.button("Đăng nhập", use_container_width=True):
            if loai_tk == "Giám thị":
                try: pass_gt = st.secrets["PASS_GT"]
                except: pass_gt = "giamthi123"
                if mat_khau == pass_gt:
                    st.session_state.update({"logged_in": True, "role": "Giám thị", "user_name": "Tổ Giám thị"})
                    st.rerun()
                else: st.error("❌ Sai mật khẩu Giám thị!")
            elif loai_tk == "Ban Giám Hiệu":
                try: pass_bgh = st.secrets["PASS_BGH"]
                except: pass_bgh = "hieutruong123"
                if mat_khau == pass_bgh:
                    st.session_state.update({"logged_in": True, "role": "BGH", "user_name": "Ban Giám Hiệu"})
                    st.rerun()
                else: st.error("❌ Sai mật khẩu Ban Giám Hiệu!")
            elif loai_tk == "Giáo viên":
                gv_match = ds_gv[ds_gv['Mã định danh'] == mat_khau.strip()]
                if not gv_match.empty:
                    st.session_state.update({"logged_in": True, "role": "Giáo viên", 
                                             "user_name": gv_match.iloc[0]['Họ tên Giáo viên'], "user_id": mat_khau.strip()})
                    st.rerun()
                else: st.error("❌ Mã định danh không tồn tại trong hệ thống!")
else:
    with st.sidebar:
        st.success(f"👤 **{st.session_state.user_name}**")
        if st.button("🚪 Đăng xuất", use_container_width=True):
            st.session_state.logged_in = False
            st.rerun()
  
        st.markdown("---")
        st.markdown(f"🟢 **Hệ thống Online**<br><small>Lần kết nối: {datetime.now().strftime('%H:%M:%S %d/%m')}</small>", unsafe_allow_html=True)

    # ==========================================
    # 6. CHỨC NĂNG GIÁM THỊ
    # ==========================================
    if st.session_state.role == "Giám thị":
        tab_gt1, tab_gt2, tab_gt3, tab_gt6, tab_gt4, tab_gt5 = st.tabs([
            "📝 Ghi nhận biến động", 
            "📤 Quản lý & Tải TKB Mới", 
            "🔎 Báo cáo Tuần", 
            "📋 Tổng hợp Công Tháng",
            "📊 Nhật ký & Điều chỉnh", 
            "➕ Thêm Giáo viên"
        ])
        
        with tab_gt1:
            st.header("Ghi nhận biến động (Nghỉ/Dạy thay)")
            col_date, _ = st.columns([1, 2])
            with col_date:
                ngay_chon = st.date_input("🗓️ Chọn ngày ghi nhận:", value=datetime.now().date())
                ngay_str = ngay_chon.strftime("%d/%m/%Y")
                tuan_hien_tai = (ngay_chon.day - 1) // 7 + 1
            
            thu_hien_tai = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"][ngay_chon.weekday()]
            
            if ngay_chon.weekday() == 6:
                st.error("🔒 HỆ THỐNG ĐÃ KHÓA SỔ. Hôm nay là Chủ Nhật, không thể cập nhật dữ liệu.")
            else:
                with st.spinner(f"Đang tải TKB_PhanCong cho ngày {ngay_str}..."):
                    df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                    df_today = df_ngoai_le[df_ngoai_le['Ngày'] == ngay_str] if not df_ngoai_le.empty else pd.DataFrame()
                    
                    tkb_phang = load_flat_tkb_by_date(ngay_chon)
                    tkb_today = tkb_phang[tkb_phang['Thứ'] == thu_hien_tai] if not tkb_phang.empty else pd.DataFrame()

                if tkb_today.empty:
                    st.warning(f"⚠️ Chưa có dữ liệu TKB_PhanCong cho ngày {ngay_str}. Hãy tải TKB lên ở tab bên cạnh.")
                else:
                    st.markdown("---")
                    col1, col2, col3 = st.columns(3)
                    classes = tkb_phang['Lớp'].unique().tolist() if not tkb_phang.empty else []
                    with col1:
                        lop = st.selectbox("Lớp", classes)
                        mon_hople = tkb_today[tkb_today['Lớp'] == lop]['Môn học'].dropna().unique().tolist()
                    if not mon_hople:
                        st.warning(f"📭 Lớp {lop} KHÔNG CÓ lịch học ngày {thu_hien_tai}.")
                    else:
                        with col1:
                            mon = st.selectbox("Môn", mon_hople)
                            tiet_hop_le = sorted([str(t) for t in tkb_today[(tkb_today['Lớp'] == lop) & (tkb_today['Môn học'] == mon)]['Tiết'].dropna().unique().tolist()])
                            tiet_list = st.multiselect("Chọn Tiết", options=tiet_hop_le, default=tiet_hop_le)

                        gv_info = tkb_today[(tkb_today['Lớp'] == lop) & (tkb_today['Môn học'] == mon)]
                        gv_goc_ten = gv_info.iloc[0]['Họ tên GV'] if not gv_info.empty else "N/A"
                        gv_goc_id = str(gv_info.iloc[0]['Mã định danh']) if not gv_info.empty else ""
                        
                        with col2:
                            st.info(f"GV Phụ trách: **{gv_goc_ten}**")
                            loai = st.selectbox("Loại", ["Nghỉ có phép", "Nghỉ không phép", "Dạy thay", "Dạy bù", "Đổi tiết", "Nghỉ Sự kiện/Thi"])
                        
                        gv_ban_list = []
                        for t in tiet_list:
                            gv_ban_list.extend(tkb_today[tkb_today['Tiết'] == str(t)]['Mã định danh'].astype(str).tolist())
                            if not df_today.empty:
                                ca_nay = df_today[df_today['Tiết'].astype(str) == str(t)]
                                gv_ban_list.extend(ca_nay['ID GV vắng'].astype(str).tolist())
                                gv_ban_list.extend(ca_nay['ID GV dạy thay'].astype(str).tolist())
                        
                        if gv_goc_id: 
                            gv_ban_list.append(gv_goc_id) 
                        gv_ban_list = list(set([x for x in gv_ban_list if x != ""]))
                        
                        # --- BỔ SUNG: LỌC GV ĐÃ NGHỈ VIỆC KHỎI DANH SÁCH DẠY THAY [cite: 12, 28, 47, 49, 52] ---
                        def check_not_resigned(row, target_date):
                            if not str(row['Ngày nghỉ việc']).strip(): return True
                            try:
                                d_resign = datetime.strptime(str(row['Ngày nghỉ việc']).strip(), "%d/%m/%Y").date()
                                return target_date < d_resign 
                            except: return True

                        df_gv_active = ds_gv[ds_gv.apply(lambda r: check_not_resigned(r, ngay_chon), axis=1)]
                        df_gv_ranh = df_gv_active[~df_gv_active['Mã định danh'].astype(str).isin(gv_ban_list)]
                        danh_sach_day_thay = ["Không"] + df_gv_ranh['Họ tên Giáo viên'].tolist()

                        with col3:
                            gv_thay_ten = st.selectbox("GV Dạy thay (Hệ thống ẩn GV đang bận)", danh_sach_day_thay)
                            gv_thay_id = str(ds_gv[ds_gv['Họ tên Giáo viên'] == gv_thay_ten]['Mã định danh'].values[0]) if gv_thay_ten != "Không" else ""
                            note = st.text_area("Ghi chú")

                        if st.button("💾 Lưu báo cáo", type="primary"):
                            if len(tiet_list) == 0:
                                st.warning("⚠️ Vui lòng chọn ít nhất 1 tiết học trước khi lưu!")
                            else:
                                trung_lap = False
                                for t in tiet_list:
                                    if not df_today.empty:
                                        ca_nay = df_today[df_today['Tiết'].astype(str) == str(t)]
                                        if not ca_nay.empty:
                                            if gv_goc_id and (gv_goc_id in ca_nay['ID GV vắng'].astype(str).values):
                                                st.error(f"⚠️ Tiết {t}: Giáo viên {gv_goc_ten} Đã báo vắng trước đó!")
                                                trung_lap = True
                                            if gv_thay_id and (gv_thay_id in ca_nay['ID GV dạy thay'].astype(str).values):
                                                st.error(f"⚠️ Tiết {t}: Giáo viên {gv_thay_ten} Đã báo dạy thay/bù trước đó!")
                                                trung_lap = True
                                
                                if not trung_lap:
                                    with st.spinner("Đang ghi nhận dữ liệu..."):
                                        rows_to_add = [[ngay_str, thu_hien_tai, t, lop, mon, loai, gv_goc_id, gv_thay_id, note] for t in tiet_list]
                                        sheet.worksheet("BaoCao_NgoaiLe").append_rows(rows_to_add)
                                        st.success(f"✅ Đã ghi nhận thành công cho ngày {ngay_str}!")
                                        st.rerun()

                    st.markdown("---")
                    with st.expander("🏖️ Khai báo Ngày Nghỉ / Sự kiện (Không tính vắng)"):
                        col_n1, col_n2 = st.columns(2)
                        pham_vi = col_n1.selectbox("Phạm vi nghỉ:", ["Toàn trường", "Chọn lớp cụ thể"])
                        lop_nghi = "ALL"
                        if pham_vi == "Chọn lớp cụ thể": 
                            lop_nghi = col_n1.selectbox("Chọn Lớp nghỉ:", classes)
                        ly_do = col_n2.text_input("Ghi chú (Tên ngày lễ, sự kiện...):")
                        if st.button("💾 Lưu Ngày Nghỉ", type="primary"):
                            sheet.worksheet("BaoCao_NgoaiLe").append_rows([[ngay_str, thu_hien_tai, "ALL", lop_nghi, "ALL", "Ngày nghỉ/Sự kiện", "ALL", "", ly_do]])
                            st.success(f"✅ Đã lưu ngày {ngay_str} là Ngày nghỉ!")

        with tab_gt2:
            st.header("Upload & Quản lý Thời Khóa Biểu")
            st.info("💡 Khi trường có TKB mới, Giám thị tải file Excel lên đây để lưu trữ theo Tuần và Tháng.")
            uploaded_file = st.file_uploader("📂 Tải lên file Thời Khóa Biểu (Excel)", type=['xlsx', 'xls'])
            if uploaded_file:
                with st.spinner("Đang AI tự động quét và nhận diện ma trận TKB..."):
                    df_raw = pd.read_excel(uploaded_file, header=None)
                    # Đọc lại DS_GV mới nhất trước khi quét TKB.
                    # Lý do: Giám thị/BGH có thể vừa bổ sung hoặc sửa Mã TKB trực tiếp trên Google Sheets.
                    load_ds_gv.clear()
                    ds_gv_upload = load_ds_gv()
                    st.write(ds_gv_upload[ds_gv_upload.astype(str).apply(lambda r: r.str.contains("Liam", case=False, na=False).any(), axis=1)])
                    df_pc, log = scan_matrix_from_dataframe(df_raw, ds_gv_upload)
                
                if df_pc.empty:
                    st.error("❌ Không thể đọc được file TKB. Vui lòng đảm bảo cấu trúc file chuẩn.")
                else:
                    st.success(f"✅ Quét thành công {len(df_pc)} tiết hợp lệ! Xem Preview bên dưới:")
                    if log:
                        with st.expander("⚠️ Có một số ô bị từ chối (Click để xem chi tiết)"):
                            for l in log: 
                                st.write(l)
                    
                    st.dataframe(normalize_dataframe_for_streamlit(df_pc), width='stretch', height=250)
                    
                    st.markdown("### Lưu TKB theo ngày áp dụng")
                    st.caption("Ngày áp dụng quyết định TKB nào được dùng khi chấm công, báo cáo và xuất lương. Ngày tải chỉ là thông tin nhật ký.")
                    col_ngay_hl, col_note_hl = st.columns([1, 2])
                    with col_ngay_hl:
                        ngay_ap_dung_tkb = st.date_input("Ngày bắt đầu áp dụng TKB:", value=datetime.now().date(), key="ngay_ap_dung_tkb_upload")
                    with col_note_hl:
                        ghi_chu_tkb = st.text_input("Ghi chú phiên bản TKB:", value="", key="ghi_chu_tkb_upload")

                    if st.button("💾 Lưu TKB theo ngày áp dụng", type="primary"):
                        with st.spinner("Đang cập nhật TKB_HieuLuc trên Google Sheets..."):
                            try:
                                ket_qua_luu = save_tkb_to_phancong(df_pc, ngay_ap_dung_tkb, ghi_chu_tkb)
                                st.success(
                                    f"🎉 Đã lưu {ket_qua_luu['so_dong_hieuluc']} dòng vào TKB_HieuLuc, "
                                    f"áp dụng từ {ngay_ap_dung_tkb.strftime('%d/%m/%Y')}."
                                )
                                if ket_qua_luu.get('so_dong_nen', 0) > 0:
                                    st.info(
                                        f"Đã tự động lưu {ket_qua_luu['so_dong_nen']} dòng TKB_PhanCong hiện có "
                                        "làm bản nền để giữ dữ liệu trước ngày áp dụng mới."
                                    )
                                if ket_qua_luu.get('so_dong_phancong', 0) > 0:
                                    st.info("TKB_PhanCong cũng đã được cập nhật vì ngày áp dụng không sau ngày hiện tại.")
                                else:
                                    st.info("TKB_PhanCong chưa ghi đè vì ngày áp dụng nằm sau ngày hiện tại; TKB mới sẽ tự có hiệu lực đúng ngày đã chọn.")
                            except Exception as e:
                                st.error(f"❌ Không cập nhật được TKB_HieuLuc: {e}")

        with tab_gt3:
            st.subheader("Báo cáo Kiểm dò chéo Sổ đầu bài")
            col_d1, col_d2 = st.columns(2)
            with col_d1: 
                start_rp = st.date_input("Từ ngày:", value=datetime.now().date() - timedelta(days=datetime.now().weekday()))
            with col_d2: 
                end_rp = st.date_input("Đến ngày:", value=start_rp + timedelta(days=6))
            
            if st.button("Tạo Báo cáo Tuần", type="primary"):
                with st.spinner("Đang tính toán số liệu tuần..."):
                    tkb_rows = []
                    cur_day = start_rp
                    danh_sach_thu = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                    while cur_day <= end_rp:
                        if cur_day.weekday() < 6:
                            thu_cur = danh_sach_thu[cur_day.weekday()]
                            tkb_ngay = load_flat_tkb_by_date(cur_day)
                            if not tkb_ngay.empty:
                                tkb_thu = tkb_ngay[tkb_ngay['Thứ'] == thu_cur].copy()
                                if not tkb_thu.empty:
                                    tkb_thu['Ngày'] = cur_day.strftime("%d/%m/%Y")
                                    tkb_rows.append(tkb_thu)
                        cur_day += timedelta(days=1)

                    if not tkb_rows:
                        st.error("❌ Không có dữ liệu TKB_PhanCong trong khoảng báo cáo.")
                    else:
                        df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                        tkb_tuan = pd.concat(tkb_rows, ignore_index=True)
                        tkb_tuan['Khối'] = "Lớp " + tkb_tuan['Lớp'].apply(extract_grade_safe)
                        rp_tkb = tkb_tuan.groupby('Khối').size().reset_index(name='Tổng TKB phải dạy')
                        
                        if not df_ngoai_le.empty:
                            df_ngoai_le['Ngày chuẩn'] = pd.to_datetime(df_ngoai_le['Ngày'], format='%d/%m/%Y', errors='coerce')
                            mask_rp = (df_ngoai_le['Ngày chuẩn'].dt.date >= start_rp) & (df_ngoai_le['Ngày chuẩn'].dt.date <= end_rp)
                            df_rp = df_ngoai_le.loc[mask_rp].copy()
                        else: 
                            df_rp = pd.DataFrame()
                        
                        if not df_rp.empty:
                            df_rp['Khối'] = "Lớp " + df_rp['Lớp'].apply(extract_grade_safe)
                            mask_su_co = ~df_rp['Loại ngoại lệ'].isin(['Ngày nghỉ/Sự kiện', 'Nghỉ Sự kiện/Thi'])
                            df_rp_loi = df_rp[mask_su_co]
                            rp_vang = df_rp_loi[df_rp_loi['ID GV vắng'].astype(str).str.strip() != ""].groupby('Khối').size().reset_index(name='Số tiết Nghỉ (Vắng)')
                            rp_thay = df_rp_loi[df_rp_loi['ID GV dạy thay'].astype(str).str.strip() != ""].groupby('Khối').size().reset_index(name='Số tiết Dạy thay')
                        else:
                            rp_vang = pd.DataFrame(columns=['Khối', 'Số tiết Nghỉ (Vắng)'])
                            rp_thay = pd.DataFrame(columns=['Khối', 'Số tiết Dạy thay'])
                        
                        rp_final = pd.merge(rp_tkb, rp_vang, on='Khối', how='left').fillna(0)
                        rp_final = pd.merge(rp_final, rp_thay, on='Khối', how='left').fillna(0)

                        for col in ['Tổng TKB phải dạy', 'Số tiết Nghỉ (Vắng)', 'Số tiết Dạy thay']:
                            if col not in rp_final.columns: rp_final[col] = 0
                            rp_final[col] = pd.to_numeric(rp_final[col], errors='coerce').fillna(0)
                        
                        rp_final['Tổng Thực Dạy'] = rp_final['Tổng TKB phải dạy'] - rp_final['Số tiết Nghỉ (Vắng)'] + rp_final['Số tiết Dạy thay']
                        
                        tong_cong = rp_final[['Tổng TKB phải dạy', 'Số tiết Nghỉ (Vắng)', 'Số tiết Dạy thay', 'Tổng Thực Dạy']].sum()
                        rp_final.loc['TOÀN TRƯỜNG'] = tong_cong
                        rp_final.at['TOÀN TRƯỜNG', 'Khối'] = "TOÀN TRƯỜNG"
                        
                        for col in ['Tổng TKB phải dạy', 'Số tiết Nghỉ (Vắng)', 'Số tiết Dạy thay', 'Tổng Thực Dạy']: 
                            rp_final[col] = pd.to_numeric(rp_final[col], errors='coerce').fillna(0).astype(int)

                        st.dataframe(normalize_dataframe_for_streamlit(rp_final), width='stretch')

        with tab_gt6:
            st.subheader("📋 Bảng Tổng hợp Ngày Giờ Công Tháng")
            col_m, col_y = st.columns(2)
            with col_m: 
                thang_bc = st.selectbox("Chọn Tháng Báo Cáo:", range(1, 13), index=datetime.now().month - 1)
            with col_y: 
                nam_bc = st.selectbox("Chọn Năm Báo Cáo:", [2024, 2025, 2026, 2027], index=2)
            
            if st.button("Tạo Bảng Tổng Hợp Tháng", type="primary"):
                with st.spinner(f"Đang quét thông minh TKB từng tuần trong tháng {thang_bc}/{nam_bc} để đối chiếu..."):
                    # --- BỔ SUNG LOGIC: ẨN GV ĐÃ NGHỈ KHỎI BẢNG TỔNG HỢP [cite: 16, 28, 47, 51, 52] ---
                    def check_active_in_month(row, m, y):
                        if not str(row['Ngày nghỉ việc']).strip(): return True
                        try:
                            d_resign = datetime.strptime(str(row['Ngày nghỉ việc']).strip(), "%d/%m/%Y").date()
                            first_day_of_report = date(y, m, 1)
                            return d_resign > first_day_of_report
                        except: return True
                    
                    ds_gv_active = ds_gv[ds_gv.apply(lambda r: check_active_in_month(r, thang_bc, nam_bc), axis=1)]

                    df_nl_all = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                    if not df_nl_all.empty:
                        df_nl_all['Ngày chuẩn'] = pd.to_datetime(df_nl_all['Ngày'], format='%d/%m/%Y', errors='coerce')
                        mask_m = (df_nl_all['Ngày chuẩn'].dt.month == thang_bc) & (df_nl_all['Ngày chuẩn'].dt.year == nam_bc)
                        df_nl_thang = df_nl_all.loc[mask_m].copy()
                    else:
                        df_nl_thang = pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])

                    dict_tkb = {}
                    danh_sach_thu_day_du = ["Thứ Hai", "Thứ Ba", "Thứ Tư", "Thứ Năm", "Thứ Sáu", "Thứ Bảy", "Chủ Nhật"]
                    nl_ngay_nghi = df_nl_thang[df_nl_thang['Loại ngoại lệ'] == 'Ngày nghỉ/Sự kiện']
                    gv_nghi_viec_map_month = tao_dict_ngay_nghi_viec(ds_gv)

                    last_day_bc = calendar.monthrange(nam_bc, thang_bc)[1]
                    for day in range(1, last_day_bc + 1):
                        ngay_iter = date(nam_bc, thang_bc, day)
                        if ngay_iter.weekday() >= 6:
                            continue
                        thu = danh_sach_thu_day_du[ngay_iter.weekday()]
                        ngay_str = f"{day:02d}/{thang_bc:02d}/{nam_bc}"

                        is_holiday = False
                        if not nl_ngay_nghi.empty:
                            match_nghi = nl_ngay_nghi[(nl_ngay_nghi['Ngày'] == ngay_str) & (nl_ngay_nghi['Lớp'] == 'ALL')]
                            if not match_nghi.empty:
                                is_holiday = True

                        if is_holiday:
                            continue

                        tkb_ngay = load_flat_tkb_by_date(ngay_iter)
                        if tkb_ngay.empty:
                            continue
                        tkb_thu = tkb_ngay[tkb_ngay['Thứ'] == thu]
                        for _, row_tkb in tkb_thu.iterrows():
                            gv_id = str(row_tkb['Mã định danh']).strip()
                            if not gv_id:
                                continue
                            if not gv_duoc_tinh_o_ngay(gv_nghi_viec_map_month.get(gv_id), ngay_iter):
                                continue
                            dict_tkb[gv_id] = dict_tkb.get(gv_id, 0) + 1
                    
                    dict_vang, dict_thay, dict_bu = {}, {}, {}
                    
                    if not df_nl_thang.empty:
                        nl_v = df_nl_thang[(df_nl_thang['ID GV vắng'].astype(str).str.strip() != '') & (df_nl_thang['Loại ngoại lệ'] != 'Nghỉ Sự kiện/Thi')]
                        for gv_id in nl_v['ID GV vắng'].astype(str).str.strip(): dict_vang[gv_id] = dict_vang.get(gv_id, 0) + 1
                        
                        nl_dt = df_nl_thang[(df_nl_thang['ID GV dạy thay'].astype(str).str.strip() != '') & (df_nl_thang['Loại ngoại lệ'] != 'Dạy bù')]
                        for gv_id in nl_dt['ID GV dạy thay'].astype(str).str.strip(): dict_thay[gv_id] = dict_thay.get(gv_id, 0) + 1
                        
                        nl_db = df_nl_thang[(df_nl_thang['ID GV dạy thay'].astype(str).str.strip() != '') & (df_nl_thang['Loại ngoại lệ'] == 'Dạy bù')]
                        for gv_id in nl_db['ID GV dạy thay'].astype(str).str.strip(): dict_bu[gv_id] = dict_bu.get(gv_id, 0) + 1

                    data_bc = []
                    # Thay đổi: Chỉ lặp qua ds_gv_active [cite: 16, 51, 52]
                    for _, row_gv in ds_gv_active.iterrows():
                        gv_id = str(row_gv['Mã định danh']).strip()
                        gv_ten = str(row_gv['Họ tên Giáo viên']).strip()
                        
                        tkb_count = dict_tkb.get(gv_id, 0)
                        vang_count = dict_vang.get(gv_id, 0)
                        thay_count = dict_thay.get(gv_id, 0)
                        bu_count = dict_bu.get(gv_id, 0)
                        
                        if tkb_count > 0 or vang_count > 0 or thay_count > 0 or bu_count > 0:
                            thuc_day = tkb_count - vang_count + thay_count + bu_count
                            data_bc.append({
                                "Mã định danh": gv_id,
                                "Họ tên Giáo viên": gv_ten,
                                "Tổng tiết phân công": tkb_count,
                                "Số tiết nghỉ": vang_count,
                                "Số tiết dạy thay": thay_count,
                                "Số tiết dạy bù": bu_count,
                                "Số tiết thực dạy": thuc_day,
                                "Ghi chú": ""
                            })
                    
                    df_bc = pd.DataFrame(data_bc)
                    if not df_bc.empty:
                        tong_cong = df_bc[['Tổng tiết phân công', 'Số tiết nghỉ', 'Số tiết dạy thay', 'Số tiết dạy bù', 'Số tiết thực dạy']].sum()
                        df_bc.loc['TỔNG TOÀN TRƯỜNG'] = tong_cong
                        df_bc.at['TỔNG TOÀN TRƯỜNG', 'Họ tên Giáo viên'] = "TỔNG TOÀN TRƯỜNG"
                        df_bc.at['TỔNG TOÀN TRƯỜNG', 'Mã định danh'] = ""
                        df_bc.at['TỔNG TOÀN TRƯỜNG', 'Ghi chú'] = ""
                        
                        for col in ['Tổng tiết phân công', 'Số tiết nghỉ', 'Số tiết dạy thay', 'Số tiết dạy bù', 'Số tiết thực dạy']:
                            df_bc[col] = pd.to_numeric(df_bc[col], errors='coerce').fillna(0).astype(int)
                        
                        df_bc = df_bc.reset_index(drop=True)
                        st.dataframe(normalize_dataframe_for_streamlit(df_bc), width='stretch')
                        
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            df_bc.to_excel(writer, index=False, sheet_name=f"Thang {thang_bc}")
                        st.download_button("📥 Tải Bảng Tổng Hợp (Excel)", data=output.getvalue(), file_name=f"TongHopCong_T{thang_bc}_{nam_bc}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    else:
                        st.info(f"📭 Không có dữ liệu TKB hay biến động nào trong tháng {thang_bc}/{nam_bc}.")


            st.markdown("---")
            st.subheader("📥 Xuất Bảng Lương Tháng")
            st.caption("Tự động lấy số tiết thực dạy từ bảng chấm công tháng và điền vào cột SL THPT trong file TH_luong.xlsx.")
            if st.button("📥 Tải Bảng Lương Tháng", type="primary"):
                with st.spinner(f"Đang tạo bảng lương tháng {thang_bc}/{nam_bc}..."):
                    dict_tkb_luong = build_tkb_by_date_cache(thang_bc, nam_bc)
                    try:
                        df_nl_luong = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                    except Exception:
                        df_nl_luong = pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])

                    excel_luong, ds_khong_khop_luong, ds_cham_cong_chua_co_luong = tao_bang_luong_thang_avm(
                        thang_bc,
                        nam_bc,
                        dict_tkb_luong,
                        df_nl_luong,
                        ds_gv
                    )

                    if excel_luong:
                        st.download_button(
                            "✅ Tải Bảng Lương Tháng",
                            data=excel_luong,
                            file_name=f"BangLuong_GV_T{thang_bc}_{nam_bc}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                        if ds_khong_khop_luong:
                            with st.expander("⚠️ Giáo viên trong file lương chưa khớp dữ liệu chấm công"):
                                for item in ds_khong_khop_luong:
                                    st.write(f"- {item}")

                        if ds_cham_cong_chua_co_luong:
                            with st.expander("⚠️ Giáo viên có chấm công nhưng chưa tìm thấy trong file lương"):
                                for item in ds_cham_cong_chua_co_luong:
                                    st.write(f"- {item}")
                    else:
                        st.error("❌ Không tạo được bảng lương tháng.")
                        for item in ds_khong_khop_luong:
                            st.warning(item)

        with tab_gt4:
            # --- CẬP NHẬT TOÀN BỘ TAB GT4: BỔ SUNG CHỨC NĂNG SỬA BIẾN ĐỘNG [cite: 17, 27, 30, 31, 32, 50, 53] ---
            st.subheader("Nhật ký Biến động & Điều chỉnh")
            with st.spinner("Đang tải dữ liệu..."):
                df_all_raw = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                if df_all_raw.empty:
                    df_filtered_gt = pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])
                else:
                    # Tạo ID_Hang để định vị dòng trên Sheets [cite: 50]
                    df_all_raw['ID_Hang'] = df_all_raw.index + 2
                    df_all_raw['Ngày chuẩn'] = pd.to_datetime(df_all_raw['Ngày'], format='%d/%m/%Y', errors='coerce')
                    min_date = df_all_raw['Ngày chuẩn'].min().date()
                    max_date = df_all_raw['Ngày chuẩn'].max().date()
                    date_range_gt = st.date_input("🗓️ Chọn khoảng thời gian xem báo cáo:", value=(min_date, max_date), key="dr_gt4")
                    
                    if isinstance(date_range_gt, tuple) and len(date_range_gt) == 2: 
                        start_date_gt, end_date_gt = date_range_gt
                    else: 
                        start_date_gt = end_date_gt = (date_range_gt[0] if isinstance(date_range_gt, tuple) else date_range_gt)
                        
                    mask_gt = (df_all_raw['Ngày chuẩn'].dt.date >= start_date_gt) & (df_all_raw['Ngày chuẩn'].dt.date <= end_date_gt)
                    df_filtered_gt = df_all_raw.loc[mask_gt].copy()

                if not df_filtered_gt.empty:
                    dict_gv_ten = pd.Series(ds_gv['Họ tên Giáo viên'].values, index=ds_gv['Mã định danh'].astype(str).str.strip()).to_dict()
                    df_filtered_gt['Giáo viên Vắng'] = df_filtered_gt['ID GV vắng'].astype(str).str.strip().map(dict_gv_ten).fillna("Không rõ")
                    df_filtered_gt['Giáo viên Dạy thay'] = df_filtered_gt['ID GV dạy thay'].astype(str).str.strip().map(dict_gv_ten).fillna("Không có")

                mask_metrics_gt = ~df_filtered_gt['Loại ngoại lệ'].isin(['Ngày nghỉ/Sự kiện', 'Nghỉ Sự kiện/Thi']) if not df_filtered_gt.empty else []
                df_metrics_gt = df_filtered_gt[mask_metrics_gt] if not df_filtered_gt.empty else pd.DataFrame()
                tong_su_co_gt = len(df_metrics_gt)
                so_ca_day_thay_gt = len(df_metrics_gt[df_metrics_gt['ID GV dạy thay'].astype(str).str.strip() != '']) if not df_metrics_gt.empty else 0
                
                col1_gt, col2_gt, col3_gt = st.columns(3)
                col1_gt.metric("Tổng tiết báo vắng", tong_su_co_gt)
                col2_gt.metric("Số tiết đã Dạy thay", so_ca_day_thay_gt)
                col3_gt.metric("Số tiết Lớp tự học", tong_su_co_gt - so_ca_day_thay_gt)
                
                if not df_filtered_gt.empty:
                    st.dataframe(normalize_dataframe_for_streamlit(df_filtered_gt[['ID_Hang', 'Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'Giáo viên Vắng', 'Giáo viên Dạy thay', 'Ghi chú']]), width='stretch')

            st.markdown("---")
            col_edit, col_del = st.columns(2)
            
            with col_edit:
                st.markdown("#### ✏️ Sửa Biến Động")
                if not df_filtered_gt.empty:
                    list_edit = [f"Dòng {r['ID_Hang']}: {r['Ngày']} - Lớp {r['Lớp']} - Tiết {r['Tiết']}" for _, r in df_filtered_gt.iterrows()]
                    sel_edit = st.selectbox("Chọn biến động cần sửa:", ["-- Chọn dòng --"] + list_edit[::-1])
                    
                    if sel_edit != "-- Chọn dòng --":
                        idx_hang = int(sel_edit.split(":")[0].replace("Dòng ", ""))
                        row_data = df_filtered_gt[df_filtered_gt['ID_Hang'] == idx_hang].iloc[0]
                        
                        with st.form("form_edit_bien_dong"):
                            list_loai = ["Nghỉ có phép", "Nghỉ không phép", "Dạy thay", "Dạy bù", "Đổi tiết", "Nghỉ Sự kiện/Thi", "Ngày nghỉ/Sự kiện"]
                            new_loai = st.selectbox("Loại mới:", list_loai, index=list_loai.index(row_data['Loại ngoại lệ']) if row_data['Loại ngoại lệ'] in list_loai else 0)
                            
                            gv_all_list = ds_gv['Họ tên Giáo viên'].tolist()
                            current_thay_name = dict_gv_ten.get(str(row_data['ID GV dạy thay']).strip(), "Không")
                            new_gv_thay_ten = st.selectbox("GV Dạy thay mới:", ["Không"] + gv_all_list, index=(gv_all_list.index(current_thay_name)+1) if current_thay_name in gv_all_list else 0)
                            new_note = st.text_area("Ghi chú mới:", value=row_data['Ghi chú'])
                            
                            if st.form_submit_button("✅ Cập nhật thay đổi"):
                                new_gv_thay_id = str(ds_gv[ds_gv['Họ tên Giáo viên'] == new_gv_thay_ten]['Mã định danh'].values[0]) if new_gv_thay_ten != "Không" else ""
                                with st.spinner("Đang cập nhật..."):
                                    ws_nl = sheet.worksheet("BaoCao_NgoaiLe")
                                    ws_nl.update_cell(idx_hang, 6, new_loai)
                                    ws_nl.update_cell(idx_hang, 8, new_gv_thay_id)
                                    ws_nl.update_cell(idx_hang, 9, new_note)
                                    st.success("✅ Đã cập nhật thành công!")
                                    st.rerun()

            with col_del:
                st.markdown("#### 🗑️ Xóa Biến Động")
                if not df_filtered_gt.empty:
                    list_del = [f"Dòng {r['ID_Hang']}: {r['Ngày']} - Tiết {r['Tiết']} ({r['Loại ngoại lệ']})" for _, r in df_filtered_gt.iterrows()]
                    sel_del = st.selectbox("Chọn biến động cần xóa:", ["-- Chọn dòng --"] + list_del[::-1])
                    if st.button("🔥 Xác nhận Xóa", type="primary"):
                        if sel_del != "-- Chọn dòng --":
                            idx_hang = int(sel_del.split(":")[0].replace("Dòng ", ""))
                            sheet.worksheet("BaoCao_NgoaiLe").delete_rows(idx_hang)
                            st.success("✅ Đã xóa thành công!")
                            st.rerun()

        with tab_gt5:
            form_them_giao_vien("giamthi")

    # ==========================================
    # 7. CHỨC NĂNG BGH & XUẤT EXCEL THÔNG MINH
    # ==========================================
    elif st.session_state.role == "BGH":
        st.header("📊 Bảng điều khiển dành cho Ban Giám Hiệu")
        
        with st.spinner("Đang tải dữ liệu toàn trường..."):
            df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
            if df_ngoai_le.empty:
                min_date, max_date = datetime.now().date(), datetime.now().date()
                df_filtered = pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])
            else:
                df_ngoai_le['Ngày chuẩn'] = pd.to_datetime(df_ngoai_le['Ngày'], format='%d/%m/%Y', errors='coerce')
                min_date = df_ngoai_le['Ngày chuẩn'].min().date()
                max_date = df_ngoai_le['Ngày chuẩn'].max().date()
                date_range = st.date_input("🗓️ Chọn khoảng thời gian xem báo cáo:", value=(min_date, max_date))
                
                if isinstance(date_range, tuple) and len(date_range) == 2: 
                    start_date, end_date = date_range
                elif isinstance(date_range, tuple) and len(date_range) == 1: 
                    start_date = end_date = date_range[0]
                else: 
                    start_date = end_date = date_range
                    
                mask = (df_ngoai_le['Ngày chuẩn'].dt.date >= start_date) & (df_ngoai_le['Ngày chuẩn'].dt.date <= end_date)
                df_filtered = df_ngoai_le.loc[mask].copy()

            dict_gv_ten = pd.Series(ds_gv['Họ tên Giáo viên'].values, index=ds_gv['Mã định danh'].astype(str).str.strip()).to_dict()
            df_filtered['Giáo viên Vắng'] = df_filtered['ID GV vắng'].astype(str).str.strip().map(dict_gv_ten).fillna("Không rõ")
            df_filtered['Giáo viên Dạy thay'] = df_filtered['ID GV dạy thay'].astype(str).str.strip().map(dict_gv_ten).fillna("Không có")

        tab1, tab2, tab3 = st.tabs(["📊 Chấm công tổng quát", "📥 Xuất EXCEL (Chấm Công Lương)", "➕ Thêm Giáo viên"])
        
        with tab1:
            st.subheader("Nhật ký Biến động Tổng quát")
            mask_metrics = ~df_filtered['Loại ngoại lệ'].isin(['Ngày nghỉ/Sự kiện', 'Nghỉ Sự kiện/Thi'])
            df_metrics = df_filtered[mask_metrics]
            tong_su_co = len(df_metrics)
            so_ca_day_thay = len(df_metrics[df_metrics['ID GV dạy thay'].astype(str).str.strip() != '']) if not df_metrics.empty else 0
            
            col1, col2, col3 = st.columns(3)
            col1.metric("Tổng tiết báo vắng", tong_su_co)
            col2.metric("Số tiết đã Dạy thay", so_ca_day_thay, delta_color="normal")
            col3.metric("Số tiết Lớp tự học", tong_su_co - so_ca_day_thay, delta_color="inverse")
            
            if not df_filtered.empty:
                st.dataframe(normalize_dataframe_for_streamlit(df_filtered[['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'Giáo viên Vắng', 'Giáo viên Dạy thay', 'Ghi chú']]), width='stretch')

        with tab2:
            st.subheader("Tạo Bảng Chấm Công Lương (Kế Thừa TKB Tự Động)")
            col_m, col_y = st.columns(2)
            with col_m: 
                thang_xuat = st.selectbox("Xuất Lương Tháng:", range(1, 13), index=datetime.now().month - 1)
            with col_y: 
                nam_xuat = st.selectbox("Năm:", [2024, 2025, 2026, 2027], index=2)
            
            dict_tkb_thang = build_tkb_by_date_cache(thang_xuat, nam_xuat)
            
            weeks = get_month_calendar(nam_xuat, thang_xuat)
            ds_gv['HienThi_BGH'] = ds_gv['Họ tên Giáo viên'] + " - ID: " + ds_gv['Mã định danh'].astype(str)
            gv_chon = st.selectbox("Chọn Giáo viên để xuất Excel:", ["-- Chọn Giáo viên --"] + ds_gv['HienThi_BGH'].tolist())
            
            st.markdown("---")
            col_ex1, col_ex2 = st.columns(2)
            df_nl_full = df_ngoai_le.copy() if not df_ngoai_le.empty else pd.DataFrame(columns=['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Loại ngoại lệ', 'ID GV vắng', 'ID GV dạy thay', 'Ghi chú'])
            gv_nghi_viec_map_export = tao_dict_ngay_nghi_viec(ds_gv)
            
            with col_ex1:
                if gv_chon != "-- Chọn Giáo viên --":
                    if st.button(f"📥 Tải Excel CÁ NHÂN ({gv_chon.split(' - ')[0]})", type="primary"):
                        with st.spinner("Đang tạo Excel..."):
                            gv_id_str = gv_chon.split(" - ID: ")[-1].strip()
                            gv_name_str = gv_chon.split(" - ID: ")[0].strip()
                            excel_data = tao_excel_mau_avm({gv_id_str: gv_name_str}, weeks, thang_xuat, nam_xuat, dict_tkb_thang, df_nl_full, is_teacher=False, gv_nghi_viec_dict=gv_nghi_viec_map_export)
                            if excel_data: 
                                st.download_button("✅ Tải File", data=excel_data, file_name=f"ChamCong_{gv_name_str}_T{thang_xuat}.xlsx")
            
            with col_ex2:
                if st.button("📥 Tải Excel TOÀN TRƯỜNG", type="primary"):
                    with st.spinner("Đang tổng hợp..."):
                        gv_dict_all = {str(row['Mã định danh']): row['Họ tên Giáo viên'] for _, row in ds_gv.iterrows() if gv_con_hieu_luc_trong_thang(gv_nghi_viec_map_export.get(str(row['Mã định danh']).strip()), thang_xuat, nam_xuat)}
                        excel_data_all = tao_excel_mau_avm(gv_dict_all, weeks, thang_xuat, nam_xuat, dict_tkb_thang, df_nl_full, is_teacher=False, gv_nghi_viec_dict=gv_nghi_viec_map_export)
                        if excel_data_all: 
                            st.download_button("✅ Tải File", data=excel_data_all, file_name=f"ChamCong_ToanTruong_T{thang_xuat}.xlsx")

        with tab3:
            form_them_giao_vien("bgh")

    # ==========================================
    # 8. CHỨC NĂNG GIÁO VIÊN (XUẤT EXCEL KHÓA)
    # ==========================================
    elif st.session_state.role == "Giáo viên":
        st.header(f"🔍 Hồ sơ đối soát của Thầy/Cô: {st.session_state.user_name}")
        
        # MỞ CỔNG TẢI FILE EXCEL TRONG 7 NGÀY ĐẦU THÁNG
        if datetime.now().day <= 7:
            today = datetime.now()
            last_m = (today.replace(day=1) - timedelta(days=1))
            t_bc, n_bc = last_m.month, last_m.year
            st.info(f"📅 Chào tháng mới! Thầy/Cô có thể tải Excel đối soát **Tháng {t_bc}/{n_bc}** (Bản bảo vệ chỉ xem).")
            
            if st.button(f"📥 Tải Excel Chấm Công Tháng {t_bc}", type="primary"):
                with st.spinner("Đang tạo file Excel bảo mật..."):
                    dict_tkb_bc = build_tkb_by_date_cache(t_bc, n_bc)
                    df_nl_full = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
                    gv_nghi_viec_map_export = tao_dict_ngay_nghi_viec(ds_gv)
                    excel_data = tao_excel_mau_avm({st.session_state.user_id: st.session_state.user_name}, 
                                                   get_month_calendar(n_bc, t_bc), t_bc, n_bc, dict_tkb_bc, df_nl_full, is_teacher=True, gv_nghi_viec_dict=gv_nghi_viec_map_export)
                    if excel_data:
                        st.download_button(label="✅ Nhấn để lưu file Excel", data=excel_data, 
                                           file_name=f"DoiSoat_{st.session_state.user_id}_T{t_bc}.xlsx", 
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        st.markdown("---")
        with st.spinner("Đang truy xuất dữ liệu cá nhân..."):
            df_ngoai_le = pd.DataFrame(sheet.worksheet("BaoCao_NgoaiLe").get_all_records())
            if not df_ngoai_le.empty:
                gv_id_str = str(st.session_state.user_id).strip()
                df_vang = df_ngoai_le[df_ngoai_le['ID GV vắng'].astype(str).str.strip() == gv_id_str].copy()
                
                if not df_vang.empty: 
                    df_vang['Vai trò'] = df_vang['Loại ngoại lệ'].apply(lambda x: "Sự kiện/Thi (Không trừ)" if x == "Nghỉ Sự kiện/Thi" else "Vắng mặt (-)")
                
                df_thay = df_ngoai_le[df_ngoai_le['ID GV dạy thay'].astype(str).str.strip() == gv_id_str].copy()
                if not df_thay.empty: 
                    df_thay['Vai trò'] = df_thay['Loại ngoại lệ'].apply(lambda x: "Dạy bù (+)" if x == "Dạy bù" else "Dạy thay (+)")
                
                df_ketqua = pd.concat([df_vang, df_thay])
                if not df_ketqua.empty:
                    st.dataframe(normalize_dataframe_for_streamlit(df_ketqua[['Ngày', 'Thứ', 'Tiết', 'Lớp', 'Môn', 'Vai trò', 'Loại ngoại lệ']]), width='stretch')
                else: 
                    st.success("🎉 Tuyệt vời! Thầy/Cô đảm bảo 100% công giảng dạy.")
            else: 
                st.info("Hệ thống hiện chưa có dữ liệu biến động nào.")